import sys
import os
from datetime import datetime
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QLabel, QLineEdit, QRadioButton, 
                             QSpinBox, QPushButton, QButtonGroup, 
                             QScrollArea, QFrame, QMessageBox, QDateEdit,
                             QFileDialog, QComboBox)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QFont
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from login import LoginWindow  # Assuming login.py is in the same directory and contains LoginWindowMain class

class FormWidget(QWidget):
    def __init__(self):
        super().__init__()
        self.excel_file_path = "./UserData/form_responses.csv"
        self.initUI()
    
    def initUI(self):
        # メインレイアウト
        main_layout = QVBoxLayout()
        
        # スクロールエリアを作成
        scroll = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)
        
        # フォームタイトル
        title = QLabel("学生情報登録フォーム")
        title.setFont(QFont("Arial", 18, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("color: #1a73e8; margin: 20px;")
        scroll_layout.addWidget(title)
        
        # 基本情報セクション
        self.add_section(scroll_layout, "基本情報")
        
        # 本名入力
        self.name_input = self.add_text_input(scroll_layout, "本名 *", "山田太郎")
        
        # メールアドレス入力
        self.email_input = self.add_text_input(scroll_layout, "メールアドレス *", "example@email.com")
        
        # 生年月日
        self.birthdate_input = self.add_date_input(scroll_layout, "生年月日 *")
        
        # 年齢選択
        self.age_input = self.add_spinbox(scroll_layout, "年齢 *", 15, 30, 20)
        
        # 性別選択（ラジオボタン）
        self.gender_group = self.add_radio_group(scroll_layout, "性別 *", ["男性", "女性", "その他"])
        
        # 現在の学年選択（コンボボックス）
        self.grade_combo = self.add_combobox(scroll_layout, "現在の学年 *", [
            "高校1年", "高校2年", "高校3年", 
            "大学1年(高専4年)", "大学2年(高専5年)", "大学3年", "大学4年",
            "大学院修士1年", "大学院修士2年", 
            "大学院博士1年", "大学院博士2年", "大学院博士3年以降",
            "その他"
        ])
        # ユーザー名・パスワード入力欄
        self.add_section(scroll_layout, "ログイン情報（新規作成）")
        self.username_input = self.add_text_input(scroll_layout, "ユーザー名 *", "半角英数字")
        self.password_input = self.add_text_input(scroll_layout, "パスワード *", "8文字以上推奨")
        self.password_input.setEchoMode(QLineEdit.Password)
        
        # ボタンセクション
        button_layout = QHBoxLayout()
        
        # 送信ボタン
        submit_btn = QPushButton("送信してCSVに保存")
        submit_btn.setStyleSheet("""
            QPushButton {
                background-color: #1a73e8;
                color: white;
                border: none;
                padding: 12px 24px;
                font-size: 16px;
                font-weight: bold;
                border-radius: 6px;
                margin: 20px;
            }
            QPushButton:hover {
                background-color: #1557b0;
            }
            QPushButton:pressed {
                background-color: #0d47a1;
            }
        """)
        submit_btn.clicked.connect(self.submit_form)
        button_layout.addWidget(submit_btn)
        
        # クリアボタン
        clear_btn = QPushButton("クリア")
        clear_btn.setStyleSheet("""
            QPushButton {
                background-color: #f8f9fa;
                color: #5f6368;
                border: 1px solid #dadce0;
                padding: 12px 24px;
                font-size: 16px;
                border-radius: 6px;
                margin: 20px;
            }
            QPushButton:hover {
                background-color: #f1f3f4;
            }
        """)
        clear_btn.clicked.connect(self.clear_form)
        button_layout.addWidget(clear_btn)
        
        # CSVファイル選択ボタン
        file_btn = QPushButton("保存先選択")
        file_btn.setStyleSheet("""
            QPushButton {
                background-color: #34a853;
                color: white;
                border: none;
                padding: 12px 24px;
                font-size: 16px;
                font-weight: bold;
                border-radius: 6px;
                margin: 20px;
            }
            QPushButton:hover {
                background-color: #2d8f45;
            }
        """)
        file_btn.clicked.connect(self.select_csv_file)
        button_layout.addWidget(file_btn)
        
        scroll_layout.addLayout(button_layout)
        
        # 現在の保存先表示
        self.file_path_label = QLabel(f"保存先: {self.excel_file_path}")
        self.file_path_label.setStyleSheet("color: #5f6368; font-size: 12px; margin: 0px 20px 20px 20px;")
        scroll_layout.addWidget(self.file_path_label)
        
        # スクロールエリアの設定
        scroll.setWidget(scroll_widget)
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border: none; }")
        
        main_layout.addWidget(scroll)
        self.setLayout(main_layout)
    
    def add_section(self, layout, title):
        """セクションタイトルを追加"""
        section_label = QLabel(title)
        section_label.setFont(QFont("Arial", 14, QFont.Bold))
        section_label.setStyleSheet("color: #202124; margin: 20px 0px 10px 0px;")
        layout.addWidget(section_label)
        
        # セクション区切り線
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        line.setStyleSheet("color: #dadce0;")
        layout.addWidget(line)
    
    def add_text_input(self, layout, label_text, placeholder=""):
        """テキスト入力フィールドを追加"""
        label = QLabel(label_text)
        label.setStyleSheet("font-weight: bold; color: #202124; margin-top: 15px;")
        layout.addWidget(label)
        
        input_field = QLineEdit()
        input_field.setPlaceholderText(placeholder)
        input_field.setStyleSheet("""
            QLineEdit {
                border: 1px solid #dadce0;
                border-radius: 4px;
                padding: 12px;
                font-size: 14px;
                margin-bottom: 10px;
            }
            QLineEdit:focus {
                border: 2px solid #1a73e8;
            }
        """)
        layout.addWidget(input_field)
        return input_field
    
    def add_spinbox(self, layout, label_text, min_val, max_val, default_val):
        """スピンボックスを追加"""
        label = QLabel(label_text)
        label.setStyleSheet("font-weight: bold; color: #202124; margin-top: 15px;")
        layout.addWidget(label)
        
        spinbox = QSpinBox()
        spinbox.setRange(min_val, max_val)
        spinbox.setValue(default_val)
        spinbox.setStyleSheet("""
            QSpinBox {
                border: 1px solid #dadce0;
                border-radius: 4px;
                padding: 12px;
                font-size: 14px;
                margin-bottom: 10px;
                min-width: 100px;
            }
            QSpinBox:focus {
                border: 2px solid #1a73e8;
            }
        """)
        layout.addWidget(spinbox)
        return spinbox
    
    def add_date_input(self, layout, label_text):
        """日付入力を追加"""
        label = QLabel(label_text)
        label.setStyleSheet("font-weight: bold; color: #202124; margin-top: 15px;")
        layout.addWidget(label)
        
        date_edit = QDateEdit()
        date_edit.setDate(QDate.currentDate().addYears(-20))  # デフォルトを20年前に設定
        date_edit.setCalendarPopup(True)
        date_edit.setStyleSheet("""
            QDateEdit {
                border: 1px solid #dadce0;
                border-radius: 4px;
                padding: 12px;
                font-size: 14px;
                margin-bottom: 10px;
                min-width: 150px;
            }
            QDateEdit:focus {
                border: 2px solid #1a73e8;
            }
        """)
        layout.addWidget(date_edit)
        return date_edit
    
    def add_radio_group(self, layout, label_text, options):
        """ラジオボタングループを追加（横並び）"""
        label = QLabel(label_text)
        label.setStyleSheet("font-weight: bold; color: #202124; margin-top: 15px;")
        layout.addWidget(label)
        
        radio_group = QButtonGroup()
        radio_layout = QHBoxLayout()
        
        for i, option in enumerate(options):
            radio_btn = QRadioButton(option)
            radio_btn.setStyleSheet("""
                QRadioButton {
                    font-size: 14px;
                    color: #202124;
                    padding: 5px;
                    margin-right: 20px;
                }
                QRadioButton::indicator {
                    width: 18px;
                    height: 18px;
                }
                QRadioButton::indicator:unchecked {
                    border: 2px solid #5f6368;
                    border-radius: 9px;
                    background-color: white;
                }
                QRadioButton::indicator:checked {
                    border: 2px solid #1a73e8;
                    border-radius: 9px;
                    background-color: #1a73e8;
                }
            """)
            if i == 0:  # デフォルトで最初の選択肢を選択
                radio_btn.setChecked(True)
            radio_group.addButton(radio_btn)
            radio_layout.addWidget(radio_btn)
        
        radio_layout.addStretch()  # 右側に余白を追加
        
        radio_widget = QWidget()
        radio_widget.setLayout(radio_layout)
        radio_widget.setStyleSheet("margin-bottom: 10px;")
        layout.addWidget(radio_widget)
        
        return radio_group
    
    def add_combobox(self, layout, label_text, options):
        """コンボボックスを追加"""
        label = QLabel(label_text)
        label.setStyleSheet("font-weight: bold; color: #202124; margin-top: 15px;")
        layout.addWidget(label)
        
        combobox = QComboBox()
        combobox.addItems(options)
        combobox.setCurrentIndex(0)  # デフォルトで最初の選択肢を選択
        combobox.setStyleSheet("""
            QComboBox {
                border: 1px solid #dadce0;
                border-radius: 4px;
                padding: 12px;
                font-size: 14px;
                margin-bottom: 10px;
                min-width: 200px;
                background-color: white;
            }
            QComboBox:focus {
                border: 2px solid #1a73e8;
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 5px solid #5f6368;
                margin-right: 5px;
            }
            QComboBox QAbstractItemView {
                border: 1px solid #dadce0;
                background-color: white;
                selection-background-color: #e8f0fe;
                selection-color: #1a73e8;
            }
        """)
        layout.addWidget(combobox)
        return combobox
    
    def select_csv_file(self):
        """CSVファイルの保存先を選択"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            "CSVファイルの保存先を選択", 
            "./UserData/form_responses.csv", 
            "CSV Files (*.csv);;All Files (*)"
        )
        
        if file_path:
            self.excel_file_path = file_path
            self.file_path_label.setText(f"保存先: {self.excel_file_path}")
    
    def save_to_csv(self, data):
        """データをCSVファイルに保存（ID対応、UTF-8 BOM付きで保存）"""
        try:
            # ヘッダー
            headers = ["ID", "登録日時", "本名", "メールアドレス", "生年月日", "年齢", "性別", "現在の学年"]
            file_exists = os.path.exists(self.excel_file_path)
            # 新規作成時はBOM付きで書き込み
            if not file_exists:
                with open(self.excel_file_path, 'w', newline='', encoding='shift-jis') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(headers)
                    writer.writerow([
                        data.get("ID", ""),
                        data["登録日時"],
                        data["本名"],
                        data["メールアドレス"],
                        data["生年月日"],
                        data["年齢"],
                        data["性別"],
                        data["現在の学年"]
                    ])
            else:
                # 既存ファイルには追記
                with open(self.excel_file_path, 'a', newline='', encoding='shift-jis') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow([
                        data.get("ID", ""),
                        data["登録日時"],
                        data["本名"],
                        data["メールアドレス"],
                        data["生年月日"],
                        data["年齢"],
                        data["性別"],
                        data["現在の学年"]
                    ])
            return True
        except Exception as e:
            QMessageBox.critical(self, "保存エラー", f"CSVファイルの保存に失敗しました:\n{str(e)}")
            return False
    
    def save_to_excel(self, data):
        """データをExcel(xlsx)ファイルに保存し、列幅を自動調整"""
        excel_path = self.excel_file_path.replace('.csv', '.xlsx')
        headers = ["登録日時", "本名", "メールアドレス", "生年月日", "年齢", "性別", "現在の学年"]
        # ファイルが存在するかチェック
        file_exists = os.path.exists(excel_path)
        if file_exists:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(headers)
        ws.append([
            data["登録日時"],
            data["本名"],
            data["メールアドレス"],
            data["生年月日"],
            data["年齢"],
            data["性別"],
            data["現在の学年"]
        ])
        # 列幅自動調整
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2
        wb.save(excel_path)
    
    def confirm_and_submit(self, data):
        """送信前に確認ダイアログを表示し、OKなら保存"""
        confirm_message = f"""
以下の内容で送信します。よろしいですか？\n\n"""
        confirm_message += f"本名: {data['本名']}\n"
        confirm_message += f"メールアドレス: {data['メールアドレス']}\n"
        confirm_message += f"生年月日: {data['生年月日']}\n"
        confirm_message += f"年齢: {data['年齢']}歳\n"
        confirm_message += f"性別: {data['性別']}\n"
        confirm_message += f"現在の学年: {data['現在の学年']}\n"
        # 追加: ユーザー名・パスワード取得
        username = self.username_input.text().strip()
        password = self.password_input.text().strip()
        confirm_message += f"\nユーザー名: {username}\nパスワード: {'*' * len(password)}"
        confirm_message += f"\n保存先: {self.excel_file_path}"
        reply = QMessageBox.question(self, "送信内容の確認", confirm_message, QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            if not username or not password:
                QMessageBox.warning(self, "入力エラー", "ユーザー名とパスワードを入力してください。")
                return
            ok, result = self.save_user_to_csv(username, password)
            if not ok:
                QMessageBox.warning(self, "登録エラー", result)
                return
            user_id = result
            data_with_id = data.copy()
            data_with_id['ID'] = user_id
            csv_result = self.save_to_csv(data_with_id)
            self.save_to_excel(data_with_id)
            if csv_result:
                result_message = f"""
データが正常に保存されました！\n\n【登録内容】\n本名: {data['本名']}\nメールアドレス: {data['メールアドレス']}\n生年月日: {data['生年月日']}\n年齢: {data['年齢']}歳\n性別: {data['性別']}\n現在の学年: {data['現在の学年']}\nユーザー名: {username}\nID: {user_id}\n\n保存先: {self.excel_file_path}
                """
                QMessageBox.information(self, "登録完了", result_message.strip())
                self.open_login_window()

    def submit_form(self):
        """フォーム送信処理"""
        # 入力値の取得
        name = self.name_input.text().strip()
        email = self.email_input.text().strip()
        birthdate = self.birthdate_input.date().toString('yyyy/MM/dd')
        age = self.age_input.value()
        
        # 性別の取得
        gender = ""
        for button in self.gender_group.buttons():
            if button.isChecked():
                gender = button.text()
                break
        
        # 学年の取得（コンボボックスから）
        grade = self.grade_combo.currentText()
        
        # 必須項目のチェック
        if not name:
            QMessageBox.warning(self, "入力エラー", "本名を入力してください。")
            self.name_input.setFocus()
            return
        
        if not email:
            QMessageBox.warning(self, "入力エラー", "メールアドレスを入力してください。")
            self.email_input.setFocus()
            return
        
        # メールアドレスの簡単な形式チェック
        if "@" not in email or "." not in email.split("@")[-1]:
            QMessageBox.warning(self, "入力エラー", "正しいメールアドレスの形式で入力してください。")
            self.email_input.setFocus()
            return
        
        # CSVに保存するデータを準備（指定された順序で）
        data = {
            "登録日時": datetime.now().strftime('%Y/%m/%d %H:%M:%S'),
            "本名": name,
            "メールアドレス": email,
            "生年月日": birthdate,
            "年齢": age,
            "性別": gender,
            "現在の学年": grade
        }
        
        # 送信前に確認
        self.confirm_and_submit(data)
    
    def clear_form(self):
        """フォームクリア処理（確認あり）"""
        reply = QMessageBox.question(self, "フォームクリア", 
                                   "全ての入力内容をクリアしますか？",
                                   QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            self.clear_form_without_confirmation()
    
    def clear_form_without_confirmation(self):
        """フォームクリア処理（確認なし）"""
        self.name_input.clear()
        self.email_input.clear()
        self.birthdate_input.setDate(QDate.currentDate().addYears(-20))
        self.age_input.setValue(20)
        self.gender_group.buttons()[0].setChecked(True)
        self.grade_combo.setCurrentIndex(0)  # コンボボックスを最初の選択肢に戻す
        self.name_input.setFocus()  # 名前欄にフォーカスを設定

    def get_next_user_id(self):
        """users.csvから最大の8桁IDを取得し、+1した値を返す"""
        import csv, os
        user_file = "./UserData/users.csv"
        max_id = 0
        if os.path.exists(user_file):
            with open(user_file, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    try:
                        id_str = row.get('ID', '0')
                        id_num = int(id_str)
                        if id_num > max_id:
                            max_id = id_num
                    except Exception:
                        pass
        return f"{max_id+1:08d}"

    def save_user_to_csv(self, username, password):
        """ユーザー情報をusers.csvに追記（重複チェックあり、IDは8桁連番）"""
        import csv, os
        user_file = "./UserData/users.csv"
        # 既存ユーザー名チェック
        if os.path.exists(user_file):
            with open(user_file, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if row.get('ユーザー名') == username:
                        return False, "このユーザー名は既に登録されています。"
        # 8桁連番ID発行
        user_id = self.get_next_user_id()
        # 新規書き込み or 追記
        file_exists = os.path.exists(user_file)
        with open(user_file, 'a', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            if not file_exists:
                writer.writerow(['ID', 'ユーザー名', 'パスワード'])
            writer.writerow([user_id, username, password])
        return True, user_id

    def open_login_window(self):
        self.login_window = LoginWindow()
        self.login_window.show()
        self.window().close()

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("学生情報登録システム")
        self.setGeometry(100, 100, 600, 800)
        
        # メインウィジェットの設定
        main_widget = FormWidget()
        self.setCentralWidget(main_widget)
        
        # ウィンドウスタイルの設定
        self.setStyleSheet("""
            QMainWindow {
                background-color: #ffffff;
            }
        """)

class FormWindowMain(QMainWindow):
    def main():
        app = QApplication(sys.argv)
    
        # アプリケーションスタイルの設定
        app.setStyle('Fusion')
    
        window = MainWindow()
        window.show()
    
        sys.exit(app.exec_())
